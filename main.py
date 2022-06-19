# Importa as bibliotecas que seram utilizadas no código.
from random import random
import yfinance
from openpyxl import Workbook
from random import randrange
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------- #
# "Rascunho" para criar uma carteira em um modelo específico:
# ------------------------------------------------------------- #

usd = yfinance.Ticker("BRL=X") # Obtém os dados do "ticket" do dólar, com relação ao real.
try:
    usd_para_brl = usd.info["regularMarketPrice"] # Seleciona o valor atual do dólar.
except:
    print("ERRO") # Caso não haja conexão com o servidor, sai do programa e imprime uma mensagem de erro. 
    exit()


# Função que converte um dado preço em dólar, para real.
def converter_brl(preco_usd):
    preco_brl = preco_usd * usd_para_brl # Multiplica o preço em dólar com o valor do dólar, em real.
    return preco_brl


# Função que retorna o preço de um determinado ativo.
def preco_ativo(ativo):
    ticker = yfinance.Ticker(ativo) # Obtém os dados do "ticket" do ativo.
    preco_ativo_usd = float(ticker.info["regularMarketPrice"]) # Seleciona o valor atual do ativo.
    preco_ativo_brl = round(converter_brl(preco_ativo_usd), 2) # Converte o preço do ativo de dólar para real
    return preco_ativo_brl


# Função que completa um dicionário da carteira, com os dados de cada ativo e o seu valor total na carteira.
def completar_carteira(carteira, tipo):
    for ativo in carteira: # Inicia um loop com cada nome de ativo na carteira.
        nome = ativo["Nome"] # Obtém a lista com os dados de cada ativo.
        if tipo == "Moedas": # Formata o nome dos tipos de moeda para ser utilizada na biblioteca.
            nome += "USD=X"
        preco = preco_ativo(nome) # Encontra o preço do ativo.
        ativo["Preço (unitário)"] = preco # Adiciona o preço do ativo na lista de dados.
        ativo["Preço (total)"] = round(preco * ativo["Quantidade"], 2) # Adiciona o preço total do ativo nos dados. 


# Dicionário com apenas os nomes dos ativos e a quantidade presente na carteira.
carteira = {"Ações": [{"Nome": "GOOGL", "Quantidade": 2}, {"Nome": "AMZN", "Quantidade": 1}, {"Nome": "AAPL", "Quantidade": 5}, 
                      {"Nome": "PETR4.SA", "Quantidade": 10}, {"Nome": "VALE", "Quantidade": 15}], 
           "Moedas": [{"Nome": "USD", "Quantidade": 150}, {"Nome": "BRL", "Quantidade": 200}, {"Nome": "EUR", "Quantidade": 100}]}

# Utiliza a função feita para completar os dados da carteira
for tipo in carteira:
    completar_carteira(carteira[tipo], tipo)

# ------------------------------------------------------------- #
# Criação do arquivo Excel, com os dados da carteira:
# ------------------------------------------------------------- #

# Função que insere os dados de uma lista em uma planilha, com cada elemento sendo inserido em uma linha diferente.
def inserir_dados(planilha, lista, linha_inicial, coluna, titulo = False):
    for linha, dado in enumerate(lista, linha_inicial):
        celula = planilha.cell(row = linha, column = coluna) # Armazena a célula em uma variável
        celula.value = dado # Insere o valor do dado na célula.
        celula.border = Border(top = Side(style='thin'), bottom = Side(style='thin')) # Insere borda superior e inferior na célula.
        if titulo: 
            celula.font = Font(bold = True) # Caso seja um título, da linha, coloca a fonte em negrito
    if titulo:
        planilha.column_dimensions[get_column_letter(coluna)].width = 15 # Altera a largura da coluna dos títulos.
    return linha # Retorna o número de linhas que foram inseridas pela função.


# Função que insere um título, com uma formatação específica, em uma planilha.
def criar_titulo(planilha, nome, linha_i, coluna_i, coluna_f):
    titulo = planilha.cell(row = linha_i, column = coluna_i) # Armazena a célula em uma variável
    titulo.value = nome # Insere o valor da célula de título.
    titulo.fill = PatternFill("solid", fgColor = "EBF1DE") # Altera a cor do preenchimento da célula.
    titulo.font = Font(bold = True, size = 14) # Aumenta a fonte e a coloca em negrito.
    titulo.alignment = Alignment(horizontal="center") # Centraliza os dados da célula.
    planilha.merge_cells(start_row = linha_i, start_column = coluna_i, end_row = linha_i, end_column = coluna_f) # Mesclar as células.


# Função que salva um planilha do excel como teste.
def salvar_arquivo_teste(workbook, max):
    nome_save = f"teste{str(randrange(max))}.xlsx" # Gera um nome para o arquivo Excel.
    workbook.save("./testes/" + nome_save) # Salva o arquivo na pasta de testes.
    print(f"Arquivo {nome_save} criado") # Imprime o nome do arquivo criado.


book = Workbook() # Inicia um workbook no Excel
tabelas = book.active # Cria uma worksheet no workbook criado

linha_tabela = 2 # Cria uma variável para armazenar a linha em que a inserção de dados começará.
coluna_tabela = 2 # Cria uma variável para armazenar a coluna em que a inserção de dados começará.
tabelas.sheet_view.showGridLines = False

for tipo in carteira:
    for num_ativo, ativo in enumerate(carteira[tipo], coluna_tabela + 1): # Loop que insere todos os dados de cada ativo na tabela.
        inserir_dados(tabelas, ativo.values(), linha_tabela + 1, num_ativo)
    criar_titulo(tabelas, tipo, linha_tabela, coluna_tabela, num_ativo) # Insere o título do tipo de ativo na tabela.    
    num_dados = inserir_dados(tabelas, carteira[tipo][0], linha_tabela + 1, coluna_tabela, titulo = True) # Insere os títulos das linhas (utilizando o primeiro ativo do tipo).
    linha_tabela += num_dados + 1 # Altera a variável linha, para criar um espaço entre as tabelas de cada tipo.

salvar_arquivo_teste(book, 1000) # Salva o arquivo teste com um número aleatório de 1 a 1000.