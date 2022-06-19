#Nesse arquivo serão gerados gráficos a partir dos dados das carteiras.

from openpyxl import load_workbook #Essa função será necessária para carregar o workbook gerado anteriormente (no branch DadosExcel).
from openpyxl.chart import BarChart, Series, Reference, LineChart #Essas funções são necessárias para construir os gráficos.
from datetime import date, timedelta, datetime #Necessário para definir as variáveis hoje e passado.
import yfinance as yf #Módulo por meio do qual é pego o histórioco dos ativos.
import pandas #Necessário na manipulação do histórico, que está em formato dataframe.

# A função a seguir um dataframe relacionado aos dados de um ativo no último ano.
def gerar(companhia, codigo):
    #Definimos o ponto de partida e chegada da nossa busca com o datetime. Depois, usamos o yfinance para buscar o histórico da companhia nesse período de tempo e em intervalos de um mês.
    hoje = date.today()
    passado = hoje - timedelta(365)
    y = yf.Ticker(codigo).history(start = passado, end = hoje, interval = "1mo")
    
    # Quando encontra uma data com dados faltantes, o yfinance já busca automaticamente outro dia próximo, anexando ambos ao dataframe.
    # Com os comandos a seguir, são eliminadas do dataframe os dias com dados faltantes, m,antendo-se um registro para cada mês.
    contador = -1
    lista = []
    tam = len(y.index)
    index = range(tam)
    y.index = index
    for linha in y.values:
        contador += 1    
        for celula in linha:
            if pandas.isna(celula) == True:
                lista.append(contador)
    x = pandas.DataFrame(data=y.drop(lista))  
    return x

wb = load_workbook(filename = 'teste189.xlsx') #Carrega o arquivo em excel.
ws = wb.active #Grava a planilha ativa.
ws_2 = wb.create_sheet() #Cria uma nova planilha.

#A seguir, iremos retirar das tabelas a informação a ser usada no gráfico e condensá-la na ws_2.
tipos = ()
valores = () #Essas tuplas armazenarão os tipos de ativo e seus valores.

#Com a estrutura a seguir, nós iteramos as linhas de interesse e a concatenamos à tupla correspondente.
for cell in ws.iter_rows(min_row=3,max_row=3,min_col=3,max_col=7,values_only=True):
    tipos = tipos + cell
for cell in ws.iter_rows(min_row=10,max_row=10,min_col=3,max_col=5,values_only=True):
    tipos = tipos + cell
for cell in ws.iter_rows(min_row=6,max_row=6,min_col=3,max_col=7,values_only=True):
    valores = valores + cell
for cell in ws.iter_rows(min_row=13,max_row=13,min_col=3,max_col=5,values_only=True):
    valores = valores + cell
#Com os comandos a seguir, as tuplas são alocadas em linhas da folha ws_2. Essas linhas serão usadas como referência no nosso gráfico.
ws_2.append(tipos)
ws_2.append(valores)

#O código a seguir é o esqueleto de um gráfico de barras.
chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Valor dos ativos"
chart1.y_axis.title = "Preço Total"
chart1.x_axis.title = "Ativos"

data = Reference(ws_2, min_col=1, min_row=1, max_row=2, max_col=8)
cats = Reference(ws_2, min_col=1, min_row=2, max_row=2, max_col=8)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4
ws.add_chart(chart1, "I2")

#Agora, um gráfico a partir dos mesmos dados, mas de barras empilhadas.
chart2 = BarChart()
chart2.type = "col"
chart2.style = 10
chart2.grouping = "stacked"
chart2.overlap=100
chart2.title = "Valor dos ativos"
chart2.y_axis.title = "Preço Total"
chart2.x_axis.title = "Ativos"

data = Reference(ws_2, min_col=1, min_row=1, max_row=2, max_col=8)
chart2.add_data(data, titles_from_data=True)
chart2.shape = 4
ws.add_chart(chart2, "I18")

#Vamos fazer um gráfico de linhas

#Cada lista a seguir será anexada à planilha como uma nova linha.
cabecalho = []
m1 = []
m2 = []
m3 = []
m4 = []
m5 = []
m6 = []
m7 = []
m8 = []
m9 = []
m10 = []
m11 = []
m12 = []
m13 = []

# O dicionário ativos contém as informações necessárias para a operação da função gerar().
ativos = {"Ambev":"ABEV3.SA", "Bancodobrasil":"BBAS3.SA", "B3":"B3SA3.SA", "Cielo":"CIEL3.SA", "Eletrobras":"ELET3.SA", "Fleury":"FLRY3.SA", "Gol":"GOLL4.SA", "Jbs":"JBSS3.SA", "Mrv":"MRVE3.SA", "Petrobras":"PETR3.SA", "Vale":"VALE3.SA"}
for companhia, codigo in ativos.items():
    x = gerar(companhia, codigo)
    
    # Para cada ativo seu nome e as informações correspondentes ao seu valos na abertura de cada dia do dataframe são adicionadas às listas. 
    cabecalho.append(companhia)
    m1.append(x.iloc[0,0])
    m2.append(x.iloc[1,0])
    m3.append(x.iloc[2,0])
    m4.append(x.iloc[3,0])
    m5.append(x.iloc[4,0])
    m6.append(x.iloc[5,0])
    m7.append(x.iloc[6,0])
    m8.append(x.iloc[7,0])
    m9.append(x.iloc[8,0])
    m10.append(x.iloc[9,0])
    m11.append(x.iloc[10,0])
    m12.append(x.iloc[11,0])
    m13.append(x.iloc[12,0])

# É criada uma tupla contendo as listas anteriores e uma lista vazia, que dará um espaço entre as informações já presentes na planilha e as que serão inclusas. Cada lista é incluída na planilha como uma linha nova.
linhas = (["","",""], cabecalho, m1, m2, m3, m4, m5, m6, m7, m8, m9, m10, m11, m12, m13)
print(linhas)
for novo in linhas:
    ws_2.append(novo)

A seguir é criado um gráfico de linha a partir dos dados coletados com o yfinance.
chart3 = LineChart()
chart3.width = 13
chart3.height = 8
chart3.title = "Evolução do valor dos ativos"
chart3.style = 13
chart3.y_axis.title = 'Valor'
chart3.x_axis.title = 'Tempo'
tanto = len(ativos)
data = Reference(ws_2, min_col=1, min_row=4, max_col=11, max_row=17)
chart3.add_data(data, titles_from_data=True)
ws.add_chart(chart3, "A15")


#Uma cópia do arquivo com as modificações feitas é criada.
wb.save("bar.xlsx")
