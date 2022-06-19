#Nesse arquivo serão gerados gráficos a partir dos dados das carteiras.

from openpyxl import load_workbook #Essa função será necessária para carregar o workbook gerado anteriormente (no branch DadosExcel).
from openpyxl.chart import BarChart, Series, Reference, LineChart #Essas funções são necessárias para construir os gráficos.
from datetime import date, timedelta
import yfinance as yf
from datetime import datetime
import pandas

def gerar(codigo):
    hoje = date.today()
    passado = hoje - timedelta(365)
    y = yf.Ticker(codigo).history(start = passado, end = hoje, interval = "1mo")
    contador = -1
    lista = []
    tam = len(y.index)
    index = range(tam)
    y.index = index
    for linha in y.values:
        contador += 1    
        for celula in linha:
            if pandas.isna(celula) == True:
                lista.append(contador)
    x = pandas.DataFrame(data=y.drop(lista))  
    return x


def agrupar_dados(carteira, ws, ws_2):
    num_acoes = len(carteira["Ações"])
    num_moedas = len(carteira["Moedas"])

    #A seguir, iremos retirar das tabelas a informação a ser usada no gráfico e condensá-la na ws_2.
    tipos = ()
    valores = () #Essas tuplas armazenarão os tipos de ativo e seus valores.

    #Com a estrutura a seguir, nós iteramos as linhas de interesse e a concatenamos à tupla correspondente.
    for cell in ws.iter_rows(min_row=3,max_row=3,min_col=3,max_col=num_acoes + 2,values_only=True):
        tipos = tipos + cell
    for cell in ws.iter_rows(min_row=10,max_row=10,min_col=3,max_col=num_moedas + 2,values_only=True):
        tipos = tipos + cell
    for cell in ws.iter_rows(min_row=6,max_row=6,min_col=3,max_col=num_acoes + 2,values_only=True):
        valores = valores + cell
    for cell in ws.iter_rows(min_row=13,max_row=13,min_col=3,max_col=num_moedas + 2,values_only=True):
        valores = valores + cell
    #Com os comandos a seguir, as tuplas são alocadas em linhas da folha ws_2. Essas linhas serão usadas como referência no nosso gráfico.
    ws_2.append(tipos)
    ws_2.append(valores)

def criar_grafico1(ativos, ws, ws_2):
    
    #O código a seguir é o esqueleto de um gráfico de barras.
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Valor dos ativos"
    chart1.y_axis.title = "Preço Total"
    chart1.x_axis.title = "Ativos"
    chart1.width = 10
    chart1.height = 12
    num_ativos = len(ativos)
    data = Reference(ws_2, min_col=1, min_row=1, max_row=2, max_col=num_ativos)
    cats = Reference(ws_2, min_col=1, min_row=2, max_row=2, max_col=num_ativos)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    ws.add_chart(chart1, "B16")

def criar_grafico2(ativos, ws, ws_2):
    
    #Agora, um gráfico a partir dos mesmos dados, mas de barras empilhadas.
    chart2 = BarChart()
    chart2.type = "col"
    chart2.style = 10
    chart2.grouping = "stacked"
    chart2.overlap=100
    chart2.title = "Valor dos ativos"
    chart2.y_axis.title = "Preço Total"
    chart2.x_axis.title = "Ativos"
    chart2.width = 10
    chart2.height = 12
    num_ativos = len(ativos)
    data = Reference(ws_2, min_col=1, min_row=1, max_row=2, max_col=num_ativos)
    chart2.add_data(data, titles_from_data=True)
    chart2.shape = 4
    ws.add_chart(chart2, "H16")

def criar_grafico3(ativos, ws, ws_2):
    
    #Vamos fazer um gráfico de linhas

    cabecalho = []
    m1 = []
    m2 = []
    m3 = []
    m4 = []
    m5 = []
    m6 = []
    m7 = []
    m8 = []
    m9 = []
    m10 = []
    m11 = []
    m12 = []
    m13 = []

    if "BRL=X" in ativos:
        ativos.remove("BRL=X")

    for codigo in ativos:
        x = gerar(codigo)
        if len(x.index) < 13:
            ativos.remove(codigo)
            continue
        cabecalho.append(codigo)
        m1.append(x.iloc[0,0])
        m2.append(x.iloc[1,0])
        m3.append(x.iloc[2,0])
        m4.append(x.iloc[3,0])
        m5.append(x.iloc[4,0])
        m6.append(x.iloc[5,0])
        m7.append(x.iloc[6,0])
        m8.append(x.iloc[7,0])
        m9.append(x.iloc[8,0])
        m10.append(x.iloc[9,0])
        m11.append(x.iloc[10,0])
        m12.append(x.iloc[11,0])
        m13.append(x.iloc[12,0])

    linhas = (["","",""], cabecalho, m1, m2, m3, m4, m5, m6, m7, m8, m9, m10, m11, m12, m13)
    for novo in linhas:
        ws_2.append(novo)
    c1 = LineChart()

    c1.width = 10
    c1.height = 12
    c1.title = "Evolução do valor dos ativos"
    c1.style = 26
    c1.y_axis.title = 'Valor'
    c1.x_axis.title = 'Tempo'

    num_ativos = len(ativos)
    data = Reference(ws_2, min_col=1, min_row=4, max_col=num_ativos, max_row=17)
    c1.add_data(data, titles_from_data=True)

    ws.add_chart(c1, "O16")

# Função que salva um planilha do excel.
def salvar_arquivo(workbook):
    nome_save = input("Digite o nome do arquivo excel que será gerado: ") # Pede um nome para o arquivo Excel.
    workbook.save(nome_save + ".xlsx") # Salva o arquivo
    print(f"Arquivo {nome_save}.xlsx criado") # Imprime o nome do arquivo criado.

    return nome_save + ".xlsx"
def criar_graficos(carteira):
    wb = load_workbook(filename = "temp.xlsx") #Carrega o arquivo em excel.
    ws = wb.active #Grava a planilha ativa.
    ws_2 = wb.create_sheet() #Cria uma nova planilha.

    ativos = []
    for lista in carteira.values():
        for dict in lista:
            ativos.append(dict['Nome'])

    agrupar_dados(carteira, ws, ws_2)
    criar_grafico1(ativos, ws, ws_2)
    criar_grafico2(ativos, ws, ws_2)
    criar_grafico3(ativos, ws, ws_2)
    #Uma cópia do arquivo com as modificações feitas é criada.
    salvar_arquivo(wb)