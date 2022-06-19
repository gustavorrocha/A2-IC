# Importa as bibliotecas que seram utilizadas no código.
from random import random
from openpyxl import Workbook
from random import randrange
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# ------------------------------------------------------------- #
# Criação do arquivo Excel, com os dados da carteira:
# ------------------------------------------------------------- #

# Função que insere os dados de uma lista em uma planilha, com cada elemento sendo inserido em uma linha diferente.
def inserir_dados(planilha, lista, linha_inicial, coluna, titulo = False):
    for linha, dado in enumerate(lista, linha_inicial):
        celula = planilha.cell(row = linha, column = coluna) # Armazena a célula em uma variável
        celula.value = dado # Insere o valor do dado na célula.
        celula.border = Border(top = Side(style='thin'), bottom = Side(style='thin')) # Insere borda superior e inferior na célula.
        if titulo: 
            celula.font = Font(bold = True) # Caso seja um título, da linha, coloca a fonte em negrito
    if titulo:
        planilha.column_dimensions[get_column_letter(coluna)].width = 15 # Altera a largura da coluna dos títulos.
    return linha # Retorna o número de linhas que foram inseridas pela função.


# Função que insere um título, com uma formatação específica, em uma planilha.
def criar_titulo(planilha, nome, linha_i, coluna_i, coluna_f):
    titulo = planilha.cell(row = linha_i, column = coluna_i) # Armazena a célula em uma variável
    titulo.value = nome # Insere o valor da célula de título.
    titulo.fill = PatternFill("solid", fgColor = "EBF1DE") # Altera a cor do preenchimento da célula.
    titulo.font = Font(bold = True, size = 14) # Aumenta a fonte e a coloca em negrito.
    titulo.alignment = Alignment(horizontal="center") # Centraliza os dados da célula.
    planilha.merge_cells(start_row = linha_i, start_column = coluna_i, end_row = linha_i, end_column = coluna_f) # Mesclar as células.


def criar_excel(carteira):
    book = Workbook() # Inicia um workbook no Excel
    tabelas = book.active # Cria uma worksheet no workbook criado

    linha_tabela = 2 # Cria uma variável para armazenar a linha em que a inserção de dados começará.
    coluna_tabela = 2 # Cria uma variável para armazenar a coluna em que a inserção de dados começará.
    tabelas.sheet_view.showGridLines = False

    for tipo in carteira:
        for num_ativo, ativo in enumerate(carteira[tipo], coluna_tabela + 1): # Loop que insere todos os dados de cada ativo na tabela.
            inserir_dados(tabelas, ativo.values(), linha_tabela + 1, num_ativo)
        criar_titulo(tabelas, tipo, linha_tabela, coluna_tabela, num_ativo) # Insere o título do tipo de ativo na tabela.    
        num_dados = inserir_dados(tabelas, carteira[tipo][0], linha_tabela + 1, coluna_tabela, titulo = True) # Insere os títulos das linhas (utilizando o primeiro ativo do tipo).
        linha_tabela += num_dados + 1 # Altera a variável linha, para criar um espaço entre as tabelas de cada tipo.

    book.save("temp.xlsx") # Salva um arquivo temporário.